from tkinter import *
import openpyxl
class N_log:
    def __init__(self):
        self.filename = ''
        self.vib_1=-1
        self.vib_2=-2
    def obrab(self, vib_1, vib_2, filename):
        self.filename = filename
        self.vib_1 = int(vib_1)
        self.vib_2 = int(vib_2)
        self.wb = openpyxl.reader.excel.load_workbook(filename=self.filename, data_only=True)
        self.sheet = self.wb.worksheets[0]
        self.rows = self.sheet.max_row
        self.cols = self.sheet.max_column
        cols_b = [chr(i + 64) for i in range(1, self.cols + 1)]
        self.tablica = [[self.sheet[j + str(i)].value for i in range(1, self.rows + 1)] for j in cols_b]
        # Создание рабочего поля из значений, указанных в рабочем листе файла
        self.rabochee_pole = [i[1:] for i in self.tablica]
        print(self.rabochee_pole)
        # Перемена активного листа
        self.sheet = self.wb.worksheets[1]
        rows_1 = self.sheet.max_row
        cols_1 = self.sheet.max_column
        cols_b_1 = [chr(i + 64) for i in range(1, cols_1 + 1)]
        tablica_1 = [[self.sheet[i + str(j)].value for j in range(1, rows_1 + 1)] for i in cols_b_1]
        term_vozr = [[int(i) for i in j.split(',')] for j in tablica_1[1][1:]]
        term_staj = [[int(i) for i in j.split(',')] for j in tablica_1[0][1:]]
        print(term_vozr, term_staj)
        m_vozrast = [self.m_func(i, self.rabochee_pole[1]) for i in term_vozr]
        m_staj = [self.m_func(i, self.rabochee_pole[2]) for i in term_staj]
        print(m_vozrast, m_staj, sep='\n')

        self.mu = m_vozrast + m_staj
        self.sheet = self.wb.worksheets[2]
        c = self.cols + 1
        r=2
        self.zapolnenie(r,c,self.mu,self.sheet)
        self.indexes = self.podhod(m_vozrast[self.vib_1],m_staj[self.vib_2])
        print(self.indexes)
        self.rezult = [self.viborka(self.indexes, self.rabochee_pole[0]),
                       self.viborka(self.indexes, self.rabochee_pole[1]),
                       self.viborka(self.indexes, self.rabochee_pole[2])]
        print(self.rezult)
        self.sheet = self.wb.worksheets[3]
        r_c=1
        r_r = 2
        self.zapolnenie(r_r,r_c,self.rezult,self.sheet)
        self.wb.save(self.filename)
    def podhod (self, pole_mu_v, pole_mu_st):
        vihod_ind=[]
        for i in range(len(pole_mu_v)):
            if max(float(pole_mu_v[i]),float(pole_mu_st[i]))==1:
                vihod_ind.append(i)
        return vihod_ind

    def viborka (self, indexes, spise):
        vih_spise = [spise[i] for i in indexes]
        return vih_spise

    def zapolnenie(self, rows, cols, spise, sheet_):
        c =cols
        for i in spise:
            r = rows
            for j in i:
                sheet_.cell(row=r, column=c).value = j
                r += 1
            c += 1



    def m_func(self,term, kandidaty):
        vihod = []
        for i in kandidaty:
            m = -1
            if i >= term[0] and i <= term[1]:
                try:
                    m = round(1 - ((term[1] - i) / (term[1] - term[0])), 2)
                except ZeroDivisionError:
                    m = 1
            elif i >= term[1] and i <= term[2]:
                m = 1
            elif i >= term[2] and i <= term[3]:
                try:
                    m = round(1 - ((i - term[2]) / (term[3] - term[2])), 2)
                except ZeroDivisionError:
                    m = 1
            else:
                m = 0
            vihod.append(m)
        return vihod



class inter(N_log):
    def __init__(self):
        self.root = Tk()
        self.root.title('Лабораторная работа №6')
        self.root.geometry('320x300')
        self.l1 = Label(self.root, text='Выберите желаемый возраст: ')
        self.l2 = Label(self.root, text='Выберите желаемый стаж: ')
        self.l3 = Label(self.root, text='Если Ваш выбор окончателен, \nнажмите кнопку')
        self.l4 = Label(self.root, text='Введите путь к файлу: ')

        self.choice_1 = IntVar(value=-1)
        self.choice_2 = IntVar(value=-2)
        self.r1 = Radiobutton(self.root, text='Молодой', variable=self.choice_1, value=0)
        self.r2 = Radiobutton(self.root, text='Средний', variable=self.choice_1, value=1)
        self.r3 = Radiobutton(self.root, text='Выше среднего', variable=self.choice_1, value=2)
        self.r4 = Radiobutton(self.root, text='Маленький', variable=self.choice_2, value=0)
        self.r5 = Radiobutton(self.root, text='Средний', variable=self.choice_2, value=1)
        self.r6 = Radiobutton(self.root, text='Большой', variable=self.choice_2, value=2)

        self.b1 = Button(self.root, text='Завершение выбора', command=self.poehali)
        self.b2 = Button(self.root, text='Выход', command=self.ubica)

        self.e1 = Entry(self.root, width = 25)

        self.l1.place(x=60, y=70)
        self.r1.place(x=20, y=90)
        self.r2.place(x=100, y=90)
        self.r3.place(x=180, y=90)
        self.l2.place(x=60, y=130)
        self.r4.place(x=20, y=150)
        self.r5.place(x=110, y=150)
        self.r6.place(x=180, y=150)
        self.l3.place(x=60, y=180)
        self.b1.place(x=80, y=220)
        self.l4.place(x=10, y=20)
        self.e1.place(x=140,y=20)
        self.b2.place(x=100, y=250)

        self.root.mainloop()
    def poehali(self):
        print(self.e1.get())
        self.obrab(self.choice_1.get(), self.choice_2.get(), self.e1.get())
    def ubica(self):
        self.root.destroy()

if __name__ == '__main__':
    a=inter()
