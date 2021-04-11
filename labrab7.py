import openpyxl
def stepen_potok(term, potok):
    vihod = []
    for i in potok:
        m = -1
        if i >= term[0] and i <= term[1]:
            try:
                m = round(1 - ((term[1] - i) / (term[1] - term[0])),2)
            except ZeroDivisionError:
                m = 1
        elif i >= term[1] and i <= term[2]:
            m = 1
        elif i >= term[2] and i <= term[3]:
            try:
                m = round(1 - ((i - term[2]) / (term[3] - term[2])),2)
            except ZeroDivisionError:
                m = 1
        else:
            m = 0
        vihod.append(m)
    return vihod
def pravila(m_term):
    inde = []
    znach =[]
    for a in m_term:
        if a!=0:
            inde.append(m_term.index(a))
            znach.append((a))
    return [inde,znach]
def classic_vrem(indes, term_vrem, znach):
    vihod = []
    for i in range(len(indes)):
        if len(indes[i])!=1:
            vihod.append(round((term_vrem[indes[i][0]][1]*znach[i][0]+term_vrem[indes[i][1]][1]*znach[i][1])/(znach[i][0]+znach[i][1]),2))
        else:
            vihod.append(term_vrem[indes[i][0]][1])

    return vihod
def vivod(m_vrem):
    vivod =[]
    for i in m_vrem:
        vivod.append(i.index(max(i)))
    return vivod
def otvet_(spise):
    itog =[]
    for i in spise:
        if i==0:
            itog.append('малая')
        elif i==1:
            itog.append('средняя')
        elif i==2:
            itog.append('большая')
    return itog
def zapolnenie(spise,sheet_,colm):
    r=2
    for i in spise:
        sheet_.cell(row=r,column=colm).value=i
        r+=1
filename = input('Введите путь к файлу: ')
wb = openpyxl.load_workbook(filename=filename)
sheet = wb.worksheets[0]
potok = [sheet.cell(i,1).value for i in range(2,sheet.max_row+1)]
print('Плотность потока ',list(potok))
sheet = wb.worksheets[1]
r1=sheet.max_row
col1=sheet.max_column
centr = int((col1+1)/2)
terms_pot = [[sheet.cell(i,j).value for j in range(2,centr+1)] for i in range(2,r1+1)]
terms_sign = [[sheet.cell(i,j).value for j in range(centr+1,col1+1)] for i in range(2,r1+1)]
print('Термы плотности потока: ',terms_pot)
print('Термы длительности сигнала: ',terms_sign)
m_pot = [stepen_potok(i,potok) for i in terms_pot]
n_m_pot = zip(m_pot[0],m_pot[1],m_pot[2])
rez = [pravila(i) for i in n_m_pot]
indes = [rez[i][0] for i in range(len(rez))]
znach = [rez[i][1] for i in range(len(rez))]
rez1 = classic_vrem(indes,terms_sign,znach)
print('Мю потока',m_pot)
print('Индексы правил: ',indes, 'Степени пригодности: ',znach,sep='\n')
print('Время на светофоре: ', rez1)
m_vrem = [stepen_potok(i, rez1) for i in terms_sign]
n_m_vrem = zip(m_vrem[0],m_vrem[1],m_vrem[2])
n_m_vrem = [list(i) for i in n_m_vrem]
print('Мю времени: ',m_vrem)
otvet = vivod(n_m_vrem)
print('Номера правил: ',otvet)
itog = otvet_(otvet)
print('Загруженность: ', itog)
sheet = wb.worksheets[0]
zapolnenie(rez1,sheet,2)
zapolnenie(itog,sheet,3)
wb.save(filename=filename)